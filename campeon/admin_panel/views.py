from django.shortcuts import render, redirect
from userauths.models import Account
from django.db.models import Q, Sum, F, DecimalField, Count
from django.core.paginator import Paginator
from django.contrib.auth.decorators import login_required
from django.views.decorators.cache import never_cache
from django.contrib.sessions.models import Session
from django.utils.timezone import now
from userauths.views import superuser_required
from products.models import Order, OrderItem, Product, Category
from django.db.models.functions import TruncDay, TruncMonth, TruncYear
from django.template.loader import render_to_string
from django.http import HttpResponse
from weasyprint import HTML
from openpyxl import Workbook
from datetime import datetime
from django.contrib import messages
import logging


logger = logging.getLogger(__name__)

# Create your views here.
@superuser_required
def user_management_search(request):
    search_user = request.GET.get("search_user")
    sort_by = request.GET.get("sort_by", "full_name")
    users = Account.objects.filter(is_superadmin=False)
    if search_user:
        users = Account.objects.filter(
            Q(full_name__icontains=search_user) | Q(email__icontains=search_user)
        )
    if sort_by in ["full_name", "email", "date_joined"]:
        users = users.order_by(sort_by)
    else:

        # users = Account.objects.filter(is_active=True)
        users = users.order_by("-full_name")

    paginator = Paginator(users, 3)

    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)
    return render(
        request,
        "admin/user_management.html",
        {"users": page_obj, "sort_by": sort_by, "page_obj": page_obj},
    )


@superuser_required
def users(request):
    users = Account.objects.filter(is_superadmin=False).order_by("-full_name")

    paginator = Paginator(users, 3)  # same as search view
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    return render(
        request,
        "admin/user_management.html",
        {"users": page_obj, "page_obj": page_obj, "sort_by": "full_name"},
    )


@superuser_required
def block_user(request, id):
    user = Account.objects.get(id=id)
    if "block_user_confimed" in request.POST:
        user.is_blocked = not user.is_blocked
        user.save()
        for session in Session.objects.filter(expire_date__gte=now()):
            data = session.get_decoded()
            # print(data)
            if data.get("_auth_user_id") == str(user.id):
                # print(data.get('_auth_user_id'))
                session.delete()
        return redirect("admin_panel:users")
    return render(request, "admin/block_user.html", {"user": user})


@superuser_required
def delete_user(request, id):
    try:
        user = Account.objects.get(id=id)
    except Account.DoesNotExist:
        messages.error(request,'User not found.')
        return redirect('admin_panel:users')
    if "delete_user_confirmed" in request.POST:
        user.delete()
        return redirect("admin_panel:users")
    return render(request, "admin/delete_user.html")


@superuser_required
def dashboard(request):

    filter_type = request.GET.get("filter", "monthly")

    orders = Order.objects.filter(payment_status="paid")

    if filter_type == "monthly":
        orders = orders.filter(created_at__year=datetime.now().year)
        

    elif filter_type == "yearly":
        pass

    order_items = OrderItem.objects.filter(order__in=orders).exclude(
        status__in=[
            "cancelled",
            "returned",
            "partially_cancelled",
            "partially_returned",
        ]
    )

    # SUMMARY

    total_sales = orders.aggregate(total=Sum("total_amount"))["total"] or 0

    total_orders = orders.count()

    total_users = Account.objects.count()

    delivered_orders = Order.objects.filter(order_status="delivered").count()

    # TOP PRODUCTS

    top_products = (
        order_items.values("variant__product__name")
        .annotate(total_sold=Sum("quantity"))
        .order_by("-total_sold")[:10]
    )

    # TOP CATEGORIES

    top_categories = (
        order_items.values("variant__product__category__name")
        .annotate(total_sold=Sum("quantity"))
        .order_by("-total_sold")[:10]
    )

    # CHART

    import calendar

    now = datetime.now()
    sales_data = []

    if filter_type == "monthly":
        # Loop through all 12 months for a complete timeline
        for month in range(1, 13):
            revenue = (
                orders.filter(created_at__month=month).aggregate(
                    total=Sum("total_amount")
                )["total"]
                or 0
            )
            sales_data.append(
                {
                    "label": calendar.month_abbr[month],  # 'Jan', 'Feb', etc.
                    "total": float(revenue),
                }
            )

    elif filter_type == "yearly":
        # Show data for the past few years up to current year
        current_year = now.year
        for year in range(current_year - 4, current_year + 1):
            revenue = (
                orders.filter(created_at__year=year).aggregate(
                    total=Sum("total_amount")
                )["total"]
                or 0
            )
            sales_data.append({"label": str(year), "total": float(revenue)})

    context = {
        "filter_type": filter_type,
        "total_sales": total_sales,
        "total_orders": total_orders,
        "total_users": total_users,
        "delivered_orders": delivered_orders,
        "top_products": top_products,
        "top_categories": top_categories,
        "sales_data": sales_data,
    }

    return render(request, "admin/dashboard.html", context)


def sales(request):
    period = request.GET.get("period", "daily")
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")

    orders = Order.objects.filter(payment_status="paid").prefetch_related("items")

    if period == "custom" and start_date and end_date:
        orders = orders.filter(
            created_at__date__gte=start_date, created_at__date__lte=end_date
        )
        group_by = TruncDay("created_at")
    elif period == "monthly":
        group_by = TruncMonth("created_at")
    elif period == "yearly":
        group_by = TruncYear("created_at")
    else:
        group_by = TruncDay("created_at")

    sales_date = (
        orders.annotate(period=group_by)
        .values("period")
        .annotate(
            order_count=Count("id"),
            sales=Sum("subtotal"),
            net_revenue=Sum("total_amount"),
        )
        .order_by("period")
    )

    sales_report = []
    for row in sales_date:
        date = row["period"]
        if period == "monthly":
            daily_orders = orders.filter(
                created_at__year=date.year, created_at__month=date.month
            )
        elif period == "yearly":
            daily_orders = orders.filter(created_at__year=date.year)
        else:
            daily_orders = orders.filter(created_at__date=date.date())
        offer_discount = sum(
            sum(item.offer_discount_amount for item in order.items.all())
            for order in daily_orders
        )
        coupon_discount = sum(
            (order.coupon_discount_value or 0) for order in daily_orders
        )
        sales = row["sales"] or 0
        net_revenue = row["net_revenue"] or 0

        sales_report.append(
            {
                "date": date,
                "orders": row["order_count"],
                "sales": sales,
                "offer_discount": offer_discount,
                "coupon_discount": coupon_discount,
                "net_revenue": net_revenue,
            }
        )
    context = {
        "sales_report": sales_report,
        "period": period,
        "start_date": start_date,
        "end_date": end_date,
        "total_orders": orders.count(),
        "total_sales": sum(order.subtotal for order in orders),
        "total_offer_discount": sum(
            sum(item.offer_discount_amount for item in order.items.all())
            for order in orders
        ),
        "total_coupon_discount": sum(
            (order.coupon_discount_value or 0) for order in orders
        ),
        "total_revenue": sum(order.total_amount for order in orders),
    }
    return render(request, "admin/sales/sales.html", context)


def sales_report_pdf(request):
    period = request.GET.get("period", "daily")
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")

    orders = Order.objects.filter(payment_status="paid").prefetch_related("items")

    if period == "custom" and start_date and end_date:
        orders = orders.filter(
            created_at__date__gte=start_date, created_at__date__lte=end_date
        )
        group_by = TruncDay("created_at")
    elif period == "monthly":
        group_by = TruncMonth("created_at")
    elif period == "yearly":
        group_by = TruncYear("created_at")
    else:
        group_by = TruncDay("created_at")

    sales_date = (
        orders.annotate(period=group_by)
        .values("period")
        .annotate(
            order_count=Count("id"),
            sales=Sum("subtotal"),
            net_revenue=Sum("total_amount"),
        )
        .order_by("period")
    )

    sales_report = []
    for row in sales_date:
        date = row["period"]

        if period == "monthly":
            daily_orders = orders.filter(
                created_at__year=date.year, created_at__month=date.month
            )
        elif period == "yearly":
            daily_orders = orders.filter(created_at__year=date.year)
        else:
            daily_orders = orders.filter(created_at__date=date.date())

        offer_discount = sum(
            sum(item.offer_discount_amount for item in order.items.all())
            for order in daily_orders
        )

        coupon_discount = sum(
            (order.coupon_discount_value or 0) for order in daily_orders
        )

        sales_report.append(
            {
                "date": date,
                "orders": row["order_count"],
                "sales": row["sales"] or 0,
                "offer_discount": offer_discount,
                "coupon_discount": coupon_discount,
                "net_revenue": row["net_revenue"] or 0,
            }
        )

    context = {
        "sales_report": sales_report,
        "total_orders": orders.count(),
        "total_sales": sum(order.subtotal for order in orders),
        "total_offer_discount": sum(
            sum(item.offer_discount_amount for item in order.items.all())
            for order in orders
        ),
        "total_coupon_discount": sum(
            (order.coupon_discount_value or 0) for order in orders
        ),
        "total_revenue": sum(order.total_amount for order in orders),
    }

    html_string = render_to_string("admin/sales/sales_report_pdf.html", context)

    html = HTML(string=html_string, base_url=request.build_absolute_uri())

    result = html.write_pdf()

    response = HttpResponse(result, content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="sales_report.pdf"'

    return response


def sales_report_excel(request):
    period = request.GET.get("period", "daily")
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")

    orders = Order.objects.filter(payment_status="paid").prefetch_related("items")

    if period == "custom" and start_date and end_date:
        orders = orders.filter(
            created_at__date__gte=start_date, created_at__date__lte=end_date
        )
        group_by = TruncDay("created_at")

    elif period == "monthly":
        group_by = TruncMonth("created_at")

    elif period == "yearly":
        group_by = TruncYear("created_at")

    else:
        group_by = TruncDay("created_at")

    sales_date = (
        orders.annotate(period=group_by)
        .values("period")
        .annotate(
            order_count=Count("id"),
            sales=Sum("subtotal"),
            net_revenue=Sum("total_amount"),
        )
        .order_by("period")
    )

    sales_report = []
    for row in sales_date:
        date = row["period"]

        if period == "monthly":
            grouped_orders = orders.filter(
                created_at__year=date.year, created_at__month=date.month
            )
        elif period == "yearly":
            grouped_orders = orders.filter(created_at__year=date.year)
        else:
            grouped_orders = orders.filter(created_at__date=date.date())

        offer_discount = sum(
            sum(item.offer_discount_amount for item in order.items.all())
            for order in grouped_orders
        )

        coupon_discount = sum(
            (order.coupon_discount_value or 0) for order in grouped_orders
        )

        sales_report.append(
            {
                "date": date,
                "orders": row["order_count"],
                "sales": row["sales"] or 0,
                "offer_discount": offer_discount,
                "coupon_discount": coupon_discount,
                "net_revenue": row["net_revenue"] or 0,
            }
        )

    # SUMMARY (same as PDF)
    total_sales = sum(r["sales"] for r in sales_report)
    total_offer_discount = sum(r["offer_discount"] for r in sales_report)
    total_coupon_discount = sum(r["coupon_discount"] for r in sales_report)
    total_revenue = sum(r["net_revenue"] for r in sales_report)

    # EXCEL FILE
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Report"

    row = 1

    # ===== TITLE =====
    ws.cell(row=row, column=1, value="SALES REPORT")
    row += 2

    # ===== SUMMARY =====
    ws.cell(row=row, column=1, value="SUMMARY")
    row += 1

    summary_data = [
        ("Total Orders", orders.count()),
        ("Total Sales", total_sales),
        ("Total Offer Discount", total_offer_discount),
        ("Total Coupon Discount", total_coupon_discount),
        ("Total Revenue", total_revenue),
    ]

    for label, value in summary_data:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=value)
        row += 1

    row += 2

    # ===== BREAKDOWN HEADER =====
    ws.cell(row=row, column=1, value="SALES BREAKDOWN")
    row += 1

    headers = [
        "Date",
        "Orders",
        "Sales",
        "Offer Discount",
        "Coupon Discount",
        "Net Revenue",
    ]

    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h)

    row += 1

    # ===== DATA =====
    for r in sales_report:
        ws.cell(row=row, column=1, value=r["date"].strftime("%d-%m-%Y"))
        ws.cell(row=row, column=2, value=r["orders"])
        ws.cell(row=row, column=3, value=r["sales"])
        ws.cell(row=row, column=4, value=r["offer_discount"])
        ws.cell(row=row, column=5, value=r["coupon_discount"])
        ws.cell(row=row, column=6, value=r["net_revenue"])
        row += 1

    # Auto width (simple)
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_len + 2

    # ===== RESPONSE =====
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="sales_report.xlsx"'

    wb.save(response)
    return response
